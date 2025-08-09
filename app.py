#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
인테리어 견적서 & 영수증 기록 생성기
Python Flask 웹 애플리케이션
"""

from flask import Flask, render_template, request, jsonify, send_file
import sqlite3
import json
import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import io

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'

# 데이터베이스 초기화
def init_db():
    conn = sqlite3.connect('estimate.db')
    cursor = conn.cursor()
    
    # 회사 정보 테이블
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS companies (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            business_number TEXT,
            address TEXT,
            ceo TEXT,
            type TEXT,
            item TEXT,
            phone TEXT,
            fax TEXT,
            manager TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # 고객 정보 테이블
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            type TEXT NOT NULL,
            name TEXT NOT NULL,
            business_number TEXT,
            address TEXT,
            ceo TEXT,
            phone TEXT,
            manager TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # 은행 계좌 테이블
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS bank_accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bank_name TEXT NOT NULL,
            account_number TEXT NOT NULL,
            account_holder TEXT NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # 견적서 테이블
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS estimates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            estimate_number TEXT,
            estimate_date DATE,
            valid_until DATE,
            company_id INTEGER,
            client_id INTEGER,
            bank_id INTEGER,
            subtotal REAL,
            tax REAL,
            total REAL,
            items TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (company_id) REFERENCES companies (id),
            FOREIGN KEY (client_id) REFERENCES clients (id),
            FOREIGN KEY (bank_id) REFERENCES bank_accounts (id)
        )
    ''')
    
    # 견적 항목 테이블
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS estimate_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            estimate_id INTEGER,
            category TEXT,
            name TEXT,
            spec TEXT,
            quantity REAL,
            price REAL,
            total REAL,
            note TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (estimate_id) REFERENCES estimates (id) ON DELETE CASCADE
        )
    ''')
    
    # 일일 기록 테이블
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS daily_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date DATE,
            site_name TEXT,
            items TEXT,
            total REAL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    conn.commit()
    conn.close()

# 한글 숫자 변환 함수
def number_to_korean(num):
    korean_numbers = ['', '일', '이', '삼', '사', '오', '육', '칠', '팔', '구']
    
    if num == 0:
        return '영원 正'
    
    result = ''
    
    # 조, 억, 만 단위로 분리
    trillion = num // 1000000000000
    hundred_million = (num % 1000000000000) // 100000000
    ten_thousand = (num % 100000000) // 10000
    remainder = num % 10000
    
    def convert_thousands(number):
        if number == 0:
            return ''
        
        str_result = ''
        thousands = number // 1000
        hundreds = (number % 1000) // 100
        tens = (number % 100) // 10
        units = number % 10
        
        if thousands > 0:
            str_result += korean_numbers[thousands] + '천'
        if hundreds > 0:
            str_result += korean_numbers[hundreds] + '백'
        if tens > 0:
            str_result += korean_numbers[tens] + '십'
        if units > 0:
            str_result += korean_numbers[units]
        
        return str_result
    
    # 조 단위 처리
    if trillion > 0:
        result += convert_thousands(trillion) + '조'
    
    # 억 단위 처리
    if hundred_million > 0:
        result += convert_thousands(hundred_million) + '억'
    
    # 만 단위 처리
    if ten_thousand > 0:
        result += convert_thousands(ten_thousand) + '만'
    
    # 나머지 천의 자리 이하 처리
    if remainder > 0:
        result += convert_thousands(remainder)
    
    return result + '원 正'

# 메인 페이지
@app.route('/')
def index():
    return render_template('index.html')

# 견적서 엑셀 생성
@app.route('/api/export_estimate_excel', methods=['POST'])
def export_estimate_excel():
    try:
        data = request.json
        
        # 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "견적서"
        
        # 스타일 정의
        header_font = Font(bold=True, size=16)
        sub_header_font = Font(bold=True, size=12)
        normal_font = Font(size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
        
        # 제목
        ws.merge_cells('A1:H1')
        ws['A1'] = '견 적 서'
        ws['A1'].font = header_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 견적 정보
        ws['A3'] = '견적번호:'
        ws['B3'] = data.get('estimate_number', '')
        ws['E3'] = '견적일자:'
        ws['F3'] = data.get('estimate_date', '')
        
        ws['A4'] = '유효기간:'
        ws['B4'] = data.get('valid_until', '')
        
        # 공급업체 정보
        ws['A6'] = '공급업체'
        ws['A6'].font = sub_header_font
        ws.merge_cells('A6:D6')
        
        company = data.get('company', {})
        ws['A7'] = f"상호: {company.get('name', '')}"
        ws['A8'] = f"사업자등록번호: {company.get('business_number', '')}"
        ws['A9'] = f"주소: {company.get('address', '')}"
        ws['A10'] = f"대표자: {company.get('ceo', '')}"
        ws['A11'] = f"전화번호: {company.get('phone', '')}"
        
        # 수요업체 정보
        ws['E6'] = '수요업체'
        ws['E6'].font = sub_header_font
        ws.merge_cells('E6:H6')
        
        client = data.get('client', {})
        ws['E7'] = f"상호: {client.get('name', '')}"
        ws['E8'] = f"사업자등록번호: {client.get('business_number', '')}"
        ws['E9'] = f"주소: {client.get('address', '')}"
        ws['E10'] = f"대표자: {client.get('ceo', '')}"
        ws['E11'] = f"전화번호: {client.get('phone', '')}"
        
        # 한글 금액 표시
        total_amount = data.get('total', 0)
        korean_amount = number_to_korean(int(total_amount))
        ws.merge_cells('A13:H13')
        ws['A13'] = f'금액: {korean_amount}'
        ws['A13'].font = Font(bold=True, size=12)
        ws['A13'].alignment = Alignment(horizontal='center')
        
        # 항목 테이블 헤더
        headers = ['공종', '품목', '규격', '단위', '수량', '단가', '공급가액', '비고']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=15, column=col, value=header)
            cell.font = sub_header_font
            cell.border = border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # 항목 데이터
        items = data.get('items', [])
        start_row = 16
        for idx, item in enumerate(items):
            row = start_row + idx
            ws.cell(row=row, column=1, value=item.get('category', '')).border = border
            ws.cell(row=row, column=2, value=item.get('name', '')).border = border
            ws.cell(row=row, column=3, value=item.get('spec', '')).border = border
            ws.cell(row=row, column=4, value=item.get('unit', 'EA')).border = border
            ws.cell(row=row, column=5, value=item.get('quantity', 0)).border = border
            ws.cell(row=row, column=6, value=item.get('price', 0)).border = border
            ws.cell(row=row, column=7, value=item.get('total', 0)).border = border
            ws.cell(row=row, column=8, value=item.get('note', '')).border = border
        
        # 합계 테이블
        summary_row = start_row + len(items) + 2
        ws.merge_cells(f'A{summary_row}:F{summary_row}')
        ws[f'A{summary_row}'] = '합계'
        
        # 공급가액, 세액, 총액 표시
        subtotal = data.get('subtotal', 0)
        tax = data.get('tax', 0)
        
        ws[f'A{summary_row + 1}'] = '공급가액'
        ws[f'B{summary_row + 1}'] = subtotal
        ws[f'C{summary_row + 1}'] = '세액'
        ws[f'D{summary_row + 1}'] = tax
        ws[f'E{summary_row + 1}'] = '합계금액'
        ws[f'F{summary_row + 1}'] = total_amount
        
        # 컬럼 너비 조정
        column_widths = [12, 25, 15, 8, 8, 15, 15, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # 파일 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{data.get('estimate_date', datetime.now().strftime('%Y-%m-%d'))}_{client.get('name', '견적서')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 영수증 기록 엑셀 생성
@app.route('/api/export_daily_excel', methods=['POST'])
def export_daily_excel():
    try:
        data = request.json
        
        # 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "영수증기록"
        
        # 스타일 정의
        header_font = Font(bold=True, size=14)
        sub_header_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        
        # 제목
        ws.merge_cells('A1:E1')
        ws['A1'] = '일일 영수증 기록 내역서'
        ws['A1'].font = header_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # 기본 정보
        ws['A3'] = '현장명:'
        ws['B3'] = data.get('site_name', '')
        ws['D3'] = '기록일자:'
        ws['E3'] = data.get('date', '')
        
        # 테이블 헤더
        headers = ['카테고리', '사용내역', '단가(원)', '금액(원)', '비고']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = sub_header_font
            cell.border = border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # 데이터 행
        items = data.get('items', [])
        total_amount = 0
        
        for idx, item in enumerate(items):
            row = 6 + idx
            ws.cell(row=row, column=1, value=item.get('category', '')).border = border
            ws.cell(row=row, column=2, value=item.get('content', '')).border = border
            ws.cell(row=row, column=3, value=item.get('rate', 0)).border = border
            ws.cell(row=row, column=4, value=item.get('amount', 0)).border = border
            ws.cell(row=row, column=5, value=item.get('note', '')).border = border
            total_amount += item.get('amount', 0)
        
        # 합계 행
        total_row = 6 + len(items) + 1
        ws.cell(row=total_row, column=3, value='합계').font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=total_amount).font = Font(bold=True)
        
        # 컬럼 너비 조정
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 25
        
        # 파일 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{data.get('date', datetime.now().strftime('%Y-%m-%d'))}_{data.get('site_name', '영수증기록')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 은행 계좌 API
@app.route('/api/bank_accounts', methods=['GET', 'POST'])
def handle_bank_accounts():
    if request.method == 'GET':
        conn = sqlite3.connect('estimate.db')
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM bank_accounts ORDER BY created_at DESC')
        accounts = cursor.fetchall()
        conn.close()
        
        columns = ['id', 'bank_name', 'account_number', 'account_holder', 'created_at']
        result = [dict(zip(columns, account)) for account in accounts]
        return jsonify(result)
    
    elif request.method == 'POST':
        data = request.json
        conn = sqlite3.connect('estimate.db')
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO bank_accounts (bank_name, account_number, account_holder)
            VALUES (?, ?, ?)
        ''', (
            data.get('bank_name', ''),
            data.get('account_number', ''),
            data.get('account_holder', '')
        ))
        account_id = cursor.lastrowid
        conn.commit()
        conn.close()
        return jsonify({'id': account_id, 'message': '계좌 정보가 저장되었습니다.'})

@app.route('/api/bank_accounts/<int:account_id>', methods=['DELETE'])
def delete_bank_account(account_id):
    conn = sqlite3.connect('estimate.db')
    cursor = conn.cursor()
    cursor.execute('DELETE FROM bank_accounts WHERE id = ?', (account_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': '계좌 정보가 삭제되었습니다.'})

# 데이터베이스 API 엔드포인트들
@app.route('/api/companies', methods=['GET', 'POST'])
def companies():
    conn = sqlite3.connect('estimate.db')
    
    if request.method == 'POST':
        data = request.json
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO companies (name, business_number, address, ceo, type, item, phone, fax, manager)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data.get('name'), data.get('business_number'), data.get('address'),
            data.get('ceo'), data.get('type'), data.get('item'),
            data.get('phone'), data.get('fax'), data.get('manager')
        ))
        conn.commit()
        company_id = cursor.lastrowid
        conn.close()
        return jsonify({'id': company_id, 'message': '회사 정보가 저장되었습니다.'})
    
    else:
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM companies ORDER BY created_at DESC')
        companies = cursor.fetchall()
        conn.close()
        return jsonify(companies)

@app.route('/api/companies/<int:company_id>', methods=['DELETE'])
def delete_company(company_id):
    conn = sqlite3.connect('estimate.db')
    cursor = conn.cursor()
    cursor.execute('DELETE FROM companies WHERE id = ?', (company_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': '회사 정보가 삭제되었습니다.'})

# 고객 정보 API
@app.route('/api/clients', methods=['GET', 'POST'])
def handle_clients():
    if request.method == 'GET':
        conn = sqlite3.connect('estimate.db')
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM clients ORDER BY created_at DESC')
        clients = cursor.fetchall()
        conn.close()
        
        # 컬럼명과 함께 결과 반환
        columns = ['id', 'type', 'name', 'business_number', 'address', 'ceo', 'phone', 'manager', 'created_at']
        result = [dict(zip(columns, client)) for client in clients]
        return jsonify(result)
    
    elif request.method == 'POST':
        data = request.json
        conn = sqlite3.connect('estimate.db')
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO clients (type, name, business_number, address, ceo, phone, manager)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            data.get('type', ''),
            data.get('name', ''),
            data.get('business_number', ''),
            data.get('address', ''),
            data.get('ceo', ''),
            data.get('contact', ''),
            data.get('manager', '')
        ))
        client_id = cursor.lastrowid
        conn.commit()
        conn.close()
        return jsonify({'id': client_id, 'message': '고객 정보가 저장되었습니다.'})

@app.route('/api/clients/<int:client_id>', methods=['DELETE'])
def delete_client(client_id):
    conn = sqlite3.connect('estimate.db')
    cursor = conn.cursor()
    cursor.execute('DELETE FROM clients WHERE id = ?', (client_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': '고객 정보가 삭제되었습니다.'})

# 견적서 데이터 API
@app.route('/api/estimates', methods=['GET', 'POST'])
def handle_estimates():
    if request.method == 'GET':
        conn = sqlite3.connect('estimate.db')
        cursor = conn.cursor()
        cursor.execute('''
            SELECT e.id, e.estimate_number, e.estimate_date, 
                   COALESCE(c.name, '') as company_name, 
                   COALESCE(cl.name, '') as client_name, 
                   e.total, e.created_at
            FROM estimates e
            LEFT JOIN companies c ON e.company_id = c.id
            LEFT JOIN clients cl ON e.client_id = cl.id
            ORDER BY e.created_at DESC
        ''')
        estimates = cursor.fetchall()
        conn.close()
        
        columns = ['id', 'estimate_number', 'estimate_date', 'company_name', 'client_name', 'total_amount', 'created_at']
        result = [dict(zip(columns, estimate)) for estimate in estimates]
        return jsonify(result)
    
    elif request.method == 'POST':
        data = request.json
        conn = sqlite3.connect('estimate.db')
        cursor = conn.cursor()
        # 회사 정보 저장 (있다면)
        company_id = None
        if 'company' in data and data['company']:
            cursor.execute('''
                INSERT INTO companies (name, business_number, address, ceo, type, item, phone, fax, manager)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                data['company'].get('name', ''),
                data['company'].get('business_number', ''),
                data['company'].get('address', ''),
                data['company'].get('ceo', ''),
                data['company'].get('type', ''),
                data['company'].get('item', ''),
                data['company'].get('phone', ''),
                data['company'].get('fax', ''),
                data['company'].get('manager', '')
            ))
            company_id = cursor.lastrowid
        
        # 고객 정보 저장 (있다면)
        client_id = None
        if 'client' in data and data['client']:
            cursor.execute('''
                INSERT INTO clients (type, name, business_number, address, ceo, phone, manager)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                data['client'].get('type', 'business'),
                data['client'].get('name', ''),
                data['client'].get('business_number', ''),
                data['client'].get('address', ''),
                data['client'].get('ceo', ''),
                data['client'].get('contact', ''),
                data['client'].get('manager', '')
            ))
            client_id = cursor.lastrowid
        
        cursor.execute('''
            INSERT INTO estimates (estimate_number, estimate_date, valid_until, company_id, client_id, subtotal, tax, total, items)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data.get('estimate_number', ''),
            data.get('estimate_date', ''),
            data.get('valid_until', ''),
            company_id,
            client_id,
            data.get('subtotal', 0),
            data.get('tax', 0),
            data.get('total', 0),
            json.dumps(data.get('items', []), ensure_ascii=False)
        ))
        estimate_id = cursor.lastrowid
        
        # 견적 항목들 저장
        items = data.get('items', [])
        for item in items:
            cursor.execute('''
                INSERT INTO estimate_items (estimate_id, category, name, spec, quantity, price, total, note)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                estimate_id,
                item.get('category', ''),
                item.get('name', ''),
                item.get('spec', ''),
                item.get('quantity', 0),
                item.get('price', 0),
                item.get('total', 0),
                item.get('note', '')
            ))
        conn.commit()
        conn.close()
        return jsonify({'id': estimate_id, 'message': '견적서가 저장되었습니다.'})

@app.route('/api/estimates/<int:estimate_id>', methods=['GET', 'DELETE'])
def handle_estimate(estimate_id):
    if request.method == 'GET':
        conn = sqlite3.connect('estimate.db')
        cursor = conn.cursor()
        cursor.execute('''
            SELECT e.*, c.name as company_name, cl.name as client_name 
            FROM estimates e
            LEFT JOIN companies c ON e.company_id = c.id
            LEFT JOIN clients cl ON e.client_id = cl.id
            WHERE e.id = ?
        ''', (estimate_id,))
        estimate = cursor.fetchone()
        
        # 견적 항목들도 함께 조회
        cursor.execute('''
            SELECT * FROM estimate_items WHERE estimate_id = ? ORDER BY id
        ''', (estimate_id,))
        estimate_items = cursor.fetchall()
        conn.close()
        
        if estimate:
            columns = ['id', 'estimate_number', 'estimate_date', 'valid_until', 'company_id', 'client_id', 'bank_id',
                      'subtotal', 'tax', 'total', 'items_json', 'created_at', 'company_name', 'client_name']
            result = dict(zip(columns, estimate))
            
            # 견적 항목들 포맷팅
            if estimate_items:
                item_columns = ['id', 'estimate_id', 'category', 'name', 'spec', 'quantity', 'price', 'total', 'note', 'created_at']
                formatted_items = []
                for item in estimate_items:
                    item_dict = dict(zip(item_columns, item))
                    formatted_items.append(item_dict)
                result['items'] = formatted_items
            else:
                # JSON 문자열에서 파싱 (fallback)
                if result['items_json']:
                    result['items'] = json.loads(result['items_json'])
                else:
                    result['items'] = []
                    
            return jsonify(result)
        else:
            return jsonify({'error': '견적서를 찾을 수 없습니다.'}), 404
    
    elif request.method == 'DELETE':
        conn = sqlite3.connect('estimate.db')
        cursor = conn.cursor()
        cursor.execute('DELETE FROM estimates WHERE id = ?', (estimate_id,))
        conn.commit()
        conn.close()
        return jsonify({'message': '견적서가 삭제되었습니다.'})

# 영수증 기록 API
@app.route('/api/daily_records', methods=['GET', 'POST'])
def handle_daily_records():
    if request.method == 'GET':
        conn = sqlite3.connect('estimate.db')
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id, date as daily_date, site_name, total, created_at
            FROM daily_records ORDER BY created_at DESC
        ''')
        records = cursor.fetchall()
        conn.close()
        
        columns = ['id', 'daily_date', 'site_name', 'total_amount', 'created_at']
        result = [dict(zip(columns, record)) for record in records]
        return jsonify(result)
    
    elif request.method == 'POST':
        data = request.json
        conn = sqlite3.connect('estimate.db')
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO daily_records (date, site_name, total, items)
            VALUES (?, ?, ?, ?)
        ''', (
            data.get('daily_date', ''),
            data.get('site_name', ''),
            data.get('total_amount', 0),
            json.dumps(data.get('items', []), ensure_ascii=False)
        ))
        record_id = cursor.lastrowid
        conn.commit()
        conn.close()
        return jsonify({'id': record_id, 'message': '영수증 기록이 저장되었습니다.'})

@app.route('/api/daily_records/<int:record_id>', methods=['GET', 'DELETE'])
def handle_daily_record(record_id):
    if request.method == 'GET':
        conn = sqlite3.connect('estimate.db')
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM daily_records WHERE id = ?', (record_id,))
        record = cursor.fetchone()
        conn.close()
        
        if record:
            columns = ['id', 'date', 'site_name', 'items', 'total', 'created_at']
            result = dict(zip(columns, record))
            # JSON 문자열을 파싱
            if result['items']:
                result['items'] = json.loads(result['items'])
            # 필드명 변경
            result['daily_date'] = result['date']
            result['total_amount'] = result['total']
            return jsonify(result)
        else:
            return jsonify({'error': '영수증 기록을 찾을 수 없습니다.'}), 404
    
    elif request.method == 'DELETE':
        conn = sqlite3.connect('estimate.db')
        cursor = conn.cursor()
        cursor.execute('DELETE FROM daily_records WHERE id = ?', (record_id,))
        conn.commit()
        conn.close()
        return jsonify({'message': '영수증 기록이 삭제되었습니다.'})

if __name__ == '__main__':
    # 데이터베이스 초기화
    init_db()
    
    # 웹 서버 실행
    print("=== 인테리어 견적서 & 영수증 기록 생성기 ===")
    print("웹 서버가 포트 5002에서 시작됩니다.")
    print("systemd로 실행 중일 때는 journalctl -u estimate-webapp -f 로 로그를 확인하세요.")
    
    # 모든 인터페이스에서 접근 가능하도록 설정
    app.run(debug=False, host='0.0.0.0', port=5002)